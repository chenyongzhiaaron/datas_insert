# coding=utf-8
import json
import random
import re
import time
import unittest
import warnings

import qiniu
from jsonpath import jsonpath
from openpyxl import load_workbook

from common.do_http import DoHttp
from common.do_mongoDB import DoMongo
from common.my_log import MyLog

sheet_name = "sheet"
remote_base_url = "https://cdn2.mdxz.pagoda.com.cn/"
base_url = "https://mdxz.kt1.pagoda.com.cn"
geocodes_host = "https://restapi.amap.com"
geocodes_method = "get"
geocodes_path = "/v3/geocode/geo?parameters"
# geocodes_key = "4390d4ccaa8df203783e767136960921"  # 6K
geocodes_key = "2c357b7ffd71ca4a932df4b81a811d63"  # 30K

address_location = r"report\logs\address_location.log"
error_geocodes = r"report\logs\error_geocodes.log"
error_location = r"report\logs\error_location.log"
error_userId = r"report\logs\error_userId.log"
no_times = r"report\logs\no_times.log"


class Execution(unittest.TestCase):

    def setUp(self) -> None:
        warnings.simplefilter('ignore', ResourceWarning)
        # 用户登录,获取 token
        user_login = DoHttp().do_http(base_url, "post",
                                      "/api/token?phone=13162583810&source=pm",
                                      r_header={"Content-Type": "application/json"}).json()
        user_token = user_login["token"]
        # 获取七牛图床 token
        self.seven_token = DoHttp().do_http(base_url, "get", "/api/uploadToken",
                                            r_header={"token": user_token}).text

    def tearDown(self) -> None:
        pass

    def test_execution(self):
        # file_name = r"datas\北京最新成熟商圈统计表.xlsx"
        # base_path = r"result\北京市成熟商圈调查明细表2020年.docx.filecontent\{}"
        # upload_image_map = base_path.format(r"word\imageMap.json")
        # image_base_path = base_path.format(r"\word") + r"\{}"
        # 统计错误结果
        count_error = []
        # 读取 excel 路径 对应 图片 jsonMap json 文件
        with open(r"D:\Program Files\datas_project\common\base_path_file_name.json", "r",
                  encoding="utf-8") as path_json:
            json_data_path = json.load(path_json)

        for file_names, base_paths in json_data_path.items():
            file_name = r"datas\{}.xlsx".format(file_names)
            base_path = r"result\{}.filecontent".format(base_paths) + r"\{}"
            upload_image_map = base_path.format(r"word\imageMap.json")
            image_base_path = base_path.format(r"\word") + r"\{}"
            # 获取 imageMap.json 中的 json 数据
            with open(upload_image_map, 'r', encoding='utf-8') as json_file:
                image_path = json.load(json_file)
            # 记录当前文件
            MyLog().my_log("." * 150, "INFO", error_location)
            MyLog().my_log("-------------------------------> {} <-------------------------------".format(file_name),
                           "INFO",
                           error_location)
            wb = load_workbook(file_name)  # 打开 excel 文件
            sheet = wb[sheet_name]
            max_row = sheet.max_row  # 获取最大行
            m_column = sheet.max_column  # 获取最大列
            """  
            first_header = []
            # 第一行标题所有值
            for i in range(2, m_column + 1):
                first_header.append(sheet.cell(2, i).value)
            """
            # 定位单元格
            for k in range(3, max_row + 1):
                data_id = sheet.cell(k, 1).value
                name = str(sheet.cell(k, 2).value).replace(" ", "").replace("/r", "").replace("/n", "")
                businessDistrictResidenceNumber = sheet.cell(k, 3).value
                realEstatePrice = sheet.cell(k, 5).value
                rent = sheet.cell(k, 6).value
                transferFee = sheet.cell(k, 7).value
                corrivals = sheet.cell(k, 8).value
                preelectionBusinessDistrict = sheet.cell(k, 9).value
                assignorContactInformation = sheet.cell(k, 10).value
                businissStatus = sheet.cell(k, 11).value
                openPlanTime_str = sheet.cell(k, 12).value
                if re.search(r"(\d+)", str(openPlanTime_str)):
                    openPlanTime = re.search(r"(\d+)", str(openPlanTime_str)).group()
                else:
                    openPlanTime = ""
                remark = sheet.cell(k, 13).value
                region = sheet.cell(k, 16).value
                phone = sheet.cell(k, 15).value
                user_name = sheet.cell(k, 14).value
                extra_city = sheet.cell(k, 17).value
                extraPois_ = sheet.cell(k, 4).value
                extraPois = []
                value_list = extraPois_.split("，")
                for each_value in value_list:
                    if re.search(r"(\d+户)", each_value):
                        number_str = re.search(r"\d+户", each_value).group()
                        number = re.search(r"(\d+)", number_str).group()
                        s_list = each_value.split(number_str)
                        try:
                            res = DoHttp().do_http(geocodes_host, geocodes_method, geocodes_path,
                                                   {"key": geocodes_key, "address": s_list[0],
                                                    "city": extra_city}).json()
                            geocodesc = jsonpath(res, "$..geocodes")
                            result_info = jsonpath(res, "$..info")[0]
                            if result_info == "OK":
                                if jsonpath(res, "$..count")[0] == "0":
                                    MyLog().my_log(
                                        "序号:{0}, 商圈名称:{1}, 小区名称:{2}, 小区定位:{3}".format(data_id, name, each_value, res),
                                        "ERROR",
                                        address_location)
                                else:
                                    pass
                            else:
                                MyLog().my_log("序号:{}, 高德:{}".format(data_id, res), "ERROR",
                                               no_times)
                        except Exception as error:
                            geocodesc = [""]
                            res = {"location": ""}
                            MyLog().my_log("高德服务查询异常; 序号:{}, 商圈名称:{}, error:{}".format(data_id, name, error), "ERROR",
                                           error_geocodes)
                        if geocodesc is not False and geocodesc[0]:
                            locate = jsonpath(res, "$..location")[0].split(",")
                            extraPois_latitude = float(locate[1])
                            extraPois_longitude = float(locate[0])
                            each_extraPois_value = {
                                "latitude": extraPois_latitude,
                                "longitude": extraPois_longitude,
                                "location": {
                                    "longitude": extraPois_longitude,
                                    "latitude": extraPois_latitude,
                                    "formattedAddress": jsonpath(res, "$..formatted_address")[0],
                                    "addressComponent": {
                                        "city": jsonpath(res, "$..city")[0],
                                        "province": jsonpath(res, "$..province")[0],
                                        "adcode": jsonpath(res, "$..adcode")[0],
                                        "district": "",
                                        "towncode": "",
                                        "country": jsonpath(res, "$..country")[0],
                                        "township": "",
                                        "citycode": ""
                                    },
                                    "customAddress": jsonpath(res, "$..formatted_address")[0]
                                },
                                "scale": number,
                                "name": s_list[0],
                                "typeName": "小区",
                                "typeUnit": "户"
                            }
                        else:
                            each_extraPois_value = {
                                "latitude": "",
                                "longitude": "",
                                "location": {
                                    "longitude": "",
                                    "latitude": "",
                                    "formattedAddress": "",
                                    "addressComponent": {
                                        "city": "",
                                        "province": "",
                                        "adcode": "",
                                        "district": "",
                                        "towncode": "",
                                        "country": "",
                                        "township": "",
                                        "citycode": ""
                                    },
                                    "customAddress": ""
                                },
                                "scale": number,
                                "name": s_list[0],
                                "typeName": "小区",
                                "typeUnit": "户"
                            }
                        extraPois.append(each_extraPois_value)
                    elif re.search(r"(\d+张)", each_value):
                        number_str = re.search(r"\d+张", each_value).group()
                        number = re.search(r"(\d+)", number_str).group()
                        s_list = each_value.split(number_str)
                        try:
                            res = DoHttp().do_http(geocodes_host, geocodes_method, geocodes_path,
                                                   {"key": geocodes_key, "address": s_list[0],
                                                    "city": extra_city}).json()
                            geocodesc = jsonpath(res, "$..geocodes")
                            result_info = jsonpath(res, "$..info")[0]
                            if result_info == "OK":
                                if jsonpath(res, "$..count")[0] == "0":
                                    MyLog().my_log(
                                        "序号:{}, 商圈名称:{}, 小区名称:{}, 小区定位:{}".format(data_id, name, each_value, res),
                                        "ERROR",
                                        address_location)
                                else:
                                    pass
                            else:
                                MyLog().my_log("序号:{}, 高德:{}".format(data_id, res), "ERROR",
                                               no_times)
                        except Exception as err:
                            geocodesc = [""]
                            res = {"location": ""}
                            MyLog().my_log("高德服务查询异常;id:{},name:{};error:{}".format(data_id, name, err), "ERROR",
                                           error_geocodes)
                        if geocodesc is not False and geocodesc[0]:
                            # print("<-------------这是 geocodes 地址对象----------------->", jsonpath(res, "$..geocodes"))
                            locate = jsonpath(res, "$..location")[0].split(",")
                            extraPois_latitude = float(locate[1])
                            extraPois_longitude = float(locate[0])
                            each_extraPois_value = {
                                "latitude": extraPois_latitude,
                                "longitude": extraPois_longitude,
                                "location": {
                                    "longitude": extraPois_longitude,
                                    "latitude": extraPois_latitude,
                                    "formattedAddress": jsonpath(res, "$..formatted_address")[0],
                                    "addressComponent": {
                                        "city": jsonpath(res, "$..city")[0],
                                        "province": jsonpath(res, "$..province")[0],
                                        "adcode": jsonpath(res, "$..adcode")[0],
                                        "district": "",
                                        "towncode": "",
                                        "country": jsonpath(res, "$..country")[0],
                                        "township": "",
                                        "citycode": jsonpath(res, "$..citycode")[0]
                                    },
                                    "customAddress": jsonpath(res, "$..formatted_address")[0]
                                },
                                "scale": number,
                                "name": s_list[0],
                                "typeName": "医院",
                                "typeUnit": "床位"
                            }
                        else:
                            each_extraPois_value = {
                                "latitude": "",
                                "longitude": "",
                                "location": {
                                    "longitude": "",
                                    "latitude": "",
                                    "formattedAddress": "",
                                    "addressComponent": {
                                        "city": "",
                                        "province": "",
                                        "adcode": "",
                                        "district": "",
                                        "towncode": "",
                                        "country": "",
                                        "township": "",
                                        "citycode": ""
                                    },
                                    "customAddress": ""
                                },
                                "scale": number,
                                "name": s_list[0],
                                "typeName": "医院",
                                "typeUnit": "床位"
                            }
                        extraPois.append(each_extraPois_value)

                # 获取图片链接及关联到预选商圈
                businessDistrictPlanImage = {}  # 平面图
                streetscapeImage = {}  # 外景图

                # 上传 7 牛图片 获取连接并关联商圈
                for image_key, image_value in image_path.items():
                    if name in image_key and "平面图" in image_key:
                        url_name = image_key + str(time.time()) + str(random.randint(1, 9) * 10)
                        file_path = image_base_path.format(image_value)
                        ret, info = qiniu.put_file(self.seven_token, url_name, file_path)
                        remote_url = remote_base_url + ret["key"]
                        businessDistrictPlanImage = {"remoteUrl": remote_url}
                    elif name in image_key and "外景图" in image_key:
                        url_name = image_key + str(time.time()) + str(random.randint(1, 9) * 10)
                        file_path = image_base_path.format(image_value)
                        ret, info = qiniu.put_file(self.seven_token, url_name, file_path)
                        remote_url = remote_base_url + ret["key"]
                        streetscapeImage = {"remoteUrl": remote_url}

                # 获取预选商圈得地址定位等详细信息
                try:
                    address_info = DoHttp().do_http(geocodes_host, geocodes_method, geocodes_path,
                                                    {"key": geocodes_key, "address": name}).json()
                    address_geocodes = jsonpath(address_info, "$..geocodes")
                    res_info = jsonpath(address_info, "$..info")[0]
                    if res_info == "OK":
                        if jsonpath(address_info, "$..count")[0] == "0":
                            MyLog().my_log("序号:{}, 商圈地址:{}, 外层定位:{}".format(data_id, name, address_info), "ERROR",
                                           address_location)
                        else:
                            pass
                    else:
                        MyLog().my_log("序号:{}, 高德:{}".format(data_id, address_info), "ERROR",
                                       no_times)
                except Exception as e:
                    address_geocodes = [""]
                    address_info = {"location": ""}
                    MyLog().my_log("高德服务查询异常; 序号:{}, name:{}, error:{}".format(data_id, name, e), "ERROR",
                                   error_geocodes)
                if address_geocodes is not False and address_geocodes[0]:
                    address_longitude = float(jsonpath(address_info, "$..location")[0].split(",")[0])
                    address_latitude = float(jsonpath(address_info, "$..location")[0].split(",")[1])
                    formatted_address = jsonpath(address_info, "$..formatted_address")[0]
                    addressComponent_city = jsonpath(address_info, "$..city")[0]
                    addressComponent_province = jsonpath(address_info, "$..province")[0]
                    addressComponent_adcode = jsonpath(address_info, "$..adcode")[0]
                    addressComponent_district = jsonpath(address_info, "$..district")[0]
                    addressComponent_citycode = ""
                    addressComponent_country = jsonpath(address_info, "$..country")[0]
                else:
                    address_longitude = ""
                    address_latitude = ""
                    formatted_address = ""
                    addressComponent_city = ""
                    addressComponent_province = ""
                    addressComponent_adcode = ""
                    addressComponent_district = ""
                    addressComponent_citycode = ""
                    addressComponent_country = ""
                # 地址详细对象
                address = {
                    "longitude": address_longitude,
                    "latitude": address_latitude,
                    "formattedAddress": formatted_address,
                    "addressComponent": {
                        "city": addressComponent_city,
                        "province": addressComponent_province,
                        "adcode": addressComponent_adcode,
                        "district": addressComponent_district,
                        "towncode": "",
                        "streetNumber": {},
                        "country": addressComponent_country,
                        "businessAreas": [],
                        "building": {},
                        "neighborhood": {},
                        "citycode": addressComponent_citycode
                    },
                    "customAddress": name
                }
                # 通过维护人电话号码查询该维护人的userId
                sql = {"phone": "{}".format(phone)}
                res_mongo = DoMongo().do_mongo(sql, "select")
                if res_mongo:
                    user_id = str(res_mongo.get("_id"))
                else:
                    user_id = ""
                    MyLog().my_log("序号:{0}, 维护人:{1}, 手机号码:{2}".format(data_id, user_name, phone), "ERROR",
                                   error_userId)
                # 创建时间与更新时间
                created_time = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.localtime(time.time()))
                updated_time = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.localtime(time.time()))
                sub_data = {"name": name, "realEstatePrice": realEstatePrice,
                            "businessDistrictResidenceNumber": businessDistrictResidenceNumber, "rent": rent,
                            "transferFee": transferFee, "corrivals": corrivals,
                            "preelectionBusinessDistrict": preelectionBusinessDistrict,
                            "assignorContactInformation": assignorContactInformation, "businissStatus": businissStatus,
                            "openPlanTime": openPlanTime, "remark": remark,
                            "region": region, "extraPois": extraPois,
                            "businessDistrictPlanImage": businessDistrictPlanImage,
                            "streetscapeImage": streetscapeImage,
                            "address": address, "city": addressComponent_city, "userId": user_id,
                            "createdAt": created_time, "updatedAt": updated_time}
                print("序号: {} 最终插入 mongo 的数据".format(data_id), json.dumps(sub_data, ensure_ascii=False))

                finally_res = jsonpath(sub_data, "$..latitude")
                finally_user_id = jsonpath(sub_data, "$..userId")
                result = []
                for i in finally_res:
                    if i == "":
                        result.append("False")
                    else:
                        result.append("True")
                for j in finally_user_id:
                    if j == "":
                        result.append("False")
                    else:
                        result.append("True")
                if "False" in result:
                    count_error.append(0)
                    MyLog().my_log(
                        "序号:{0},名称:{1},小区户数:{2},城市:{3},号码:{4},维护人:{5},userId:{6}".format(data_id, name, value_list,
                                                                                         extra_city,
                                                                                         phone, user_name, user_id),
                        "ERROR",
                        error_location)
                else:
                    # 符合条件的数据插入 mongodb
                    DoMongo().do_mongo(sub_data)
            MyLog().my_log("-------------------------------> {} <-------------------------------".format(file_name),
                           "INFO",
                           error_location)


# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    unittest.main()
